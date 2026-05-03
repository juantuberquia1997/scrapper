"""
Supermu Discount Tracker
Fetches all products via /products.json pagination, exports an Excel report,
and sends it by email.

Usage:
  python site_prober.py   # run the tracker
"""
import time
import random
import smtplib
import os
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

import requests
import openpyxl
from tqdm import tqdm
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# ---------------------------------------------------------------------------
# Email config
# ---------------------------------------------------------------------------
ENABLE_EMAIL    = bool(os.getenv("SUPERMU_EMAIL"))
SMTP_SERVER     = "smtp.gmail.com"
SMTP_PORT       = 587
SENDER_EMAIL    = os.getenv("SUPERMU_EMAIL", "")
SENDER_PASSWORD = os.getenv("SUPERMU_PASSWORD", "")
RECIPIENT_EMAIL = os.getenv("SUPERMU_RECIPIENT", "")

# ---------------------------------------------------------------------------
# Frequent products — loaded from products_list.txt (same folder as script)
# Edit that file to add/remove products without touching this script.
# ---------------------------------------------------------------------------
def _load_frequent_products() -> list[str]:
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "products_list.txt")
    products = []
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line.startswith('"') and line.endswith('",'):
                    products.append(line[1:-2].strip())
                elif line.startswith('"') and line.endswith('"'):
                    products.append(line[1:-1].strip())
    except FileNotFoundError:
        print("  [WARN] No se encontró products_list.txt — columna Frecuente desactivada")
    return products

_FREQUENT_KEYS = [name.upper() for name in _load_frequent_products()]

def is_frequent(title: str) -> bool:
    t = title.upper()
    return any(t.startswith(key) for key in _FREQUENT_KEYS)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BASE_URL    = "https://supermu.com"
RETRY_WAIT  = 15
MAX_RETRIES = 3
PAGE_DELAY  = (0.5, 1.2)  # seconds between pages — be gentle with rate limits

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/91.0.4472.124 Safari/537.36"
    )
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fmt_cop(value: float | None) -> str:
    return f"${value:,.0f}" if value is not None else ""


# ---------------------------------------------------------------------------
# Scraper — paginates /products.json until empty page
# ---------------------------------------------------------------------------

def _fetch_products_page(page: int) -> list[dict]:
    url = f"{BASE_URL}/products.json?limit=250&page={page}"
    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=15)
            if resp.status_code == 429:
                wait = RETRY_WAIT * (attempt + 1)
                tqdm.write(f"  [429] Rate limit — esperando {wait}s...")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            return resp.json().get("products", [])
        except Exception as e:
            tqdm.write(f"  [ERROR] página {page}: {e}")
            return []
    return []


def scrape_all_products() -> list[dict]:
    results = []
    seen_ids: set[int] = set()
    page = 1

    with tqdm(desc="  Páginas", unit="pág", ncols=70) as bar:
        while True:
            products = _fetch_products_page(page)
            if not products:
                break

            for p in products:
                pid = p.get("id")
                if pid in seen_ids:
                    continue
                seen_ids.add(pid)

                variant  = p.get("variants", [{}])[0]
                price    = float(variant.get("price") or 0)
                compare  = variant.get("compare_at_price")
                compare  = float(compare) if compare else None
                on_sale  = bool(compare and compare > price)
                in_stock = bool(variant.get("available", True))
                title    = p.get("title", "")
                category = p.get("product_type", "")

                results.append({
                    "title":            title,
                    "url":              f"{BASE_URL}/products/{p.get('handle', '')}",
                    "category":         category,
                    "found":            True,
                    "in_stock":         in_stock,
                    "frequent":         is_frequent(title),
                    "has_discount":     on_sale,
                    "original_price":   compare if on_sale else price,
                    "discounted_price": price if on_sale else None,
                    "savings_cop":      round(compare - price) if on_sale else None,
                    "savings_pct":      round((compare - price) / compare * 100, 1) if on_sale else None,
                    "discount_label":   f"Ahorro {round((compare - price) / compare * 100, 1)}%" if on_sale else "",
                })

            bar.update(1)
            bar.set_postfix(productos=len(results))

            if len(products) < 250:
                break
            page += 1
            time.sleep(random.uniform(*PAGE_DELAY))

    return results


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

C_GREEN_DARK  = "1F5C2E"
C_GREEN_LIGHT = "E8F5E9"
C_ORANGE      = "E65100"
C_YELLOW      = "FFF9C4"
C_WHITE       = "FFFFFF"
C_GRAY        = "F5F5F5"
C_RED_LIGHT   = "FFEBEE"
C_BLUE_LIGHT  = "E3F2FD"
C_BLUE_DARK   = "1565C0"


def _hcell(ws, row, col, value, bg=C_GREEN_DARK, fg=C_WHITE, size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, color=fg, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_excel(results: list[dict], filename: str) -> None:
    wb = openpyxl.Workbook()

    disc_instock = [r for r in results if r["has_discount"] and r.get("in_stock", True)]
    disc_nostock = [r for r in results if r["has_discount"] and not r.get("in_stock", True)]
    not_found    = [r for r in results if not r["found"]]
    no_discount  = [r for r in results if r["found"] and not r["has_discount"]]

    COLS = ["Frecuente", "Producto", "Categoría", "Disponible", "Precio Original (COP)",
            "Precio con Descuento (COP)", "Ahorro (COP)", "Ahorro (%)", "Etiqueta Promocion", "URL"]
    WIDTHS = [11, 55, 25, 12, 24, 26, 18, 12, 22, 65]

    def _write_discount_rows(ws, rows):
        for ri, r in enumerate(sorted(rows, key=lambda x: x["savings_pct"] or 0, reverse=True), start=2):
            in_stock = r.get("in_stock", True)
            frequent = r.get("frequent", False)
            bg = C_GREEN_LIGHT if (in_stock and ri % 2 == 0) else \
                 C_WHITE       if in_stock else C_BLUE_LIGHT
            row = ["Sí" if frequent else "", r["title"], r["category"],
                   "Sí" if in_stock else "No",
                   r["original_price"], r["discounted_price"],
                   r["savings_cop"], r["savings_pct"], r["discount_label"], r["url"]]
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(vertical="center", horizontal="center" if ci == 1 else "left")
                if ci in (5, 6, 7) and isinstance(val, (int, float)):
                    cell.number_format = '"$"#,##0'
                if ci == 8 and isinstance(val, (int, float)):
                    cell.number_format = '0.0"%"'
                    if val >= 20:
                        cell.font = Font(bold=True, color=C_ORANGE)
                if ci == 4 and not in_stock:
                    cell.font = Font(bold=True, color=C_BLUE_DARK)

    # ── Sheet 1: Descuento CON stock ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Con Descuento"
    for c, h in enumerate(COLS, 1):
        _hcell(ws1, 1, c, h)
    _col_widths(ws1, WIDTHS)
    ws1.row_dimensions[1].height = 30
    ws1.freeze_panes = "A2"
    _write_discount_rows(ws1, disc_instock)

    # ── Sheet 2: Descuento SIN stock ────────────────────────────────────────
    ws2 = wb.create_sheet("Descuento Sin Stock")
    for c, h in enumerate(COLS, 1):
        _hcell(ws2, 1, c, h, bg=C_BLUE_DARK)
    _col_widths(ws2, WIDTHS)
    ws2.row_dimensions[1].height = 30
    ws2.freeze_panes = "A2"
    _write_discount_rows(ws2, disc_nostock)

    # ── Sheet 3: All products ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Todos los Resultados")
    h3 = ["#", "Frecuente", "Producto", "Categoría", "Disponible", "Estado",
          "Precio Original (COP)", "Precio Desc. (COP)", "Ahorro (%)", "URL"]
    w3 = [5, 11, 55, 25, 12, 18, 24, 24, 12, 65]
    for c, h in enumerate(h3, 1):
        _hcell(ws3, 1, c, h)
    _col_widths(ws3, w3)
    ws3.row_dimensions[1].height = 30
    ws3.freeze_panes = "B2"

    for ri, r in enumerate(results, start=2):
        in_stock = r.get("in_stock", True)
        if not r["found"]:
            status, bg = "Error al leer", C_RED_LIGHT
        elif r["has_discount"] and in_stock:
            status, bg = "DESCUENTO", C_YELLOW
        elif r["has_discount"] and not in_stock:
            status, bg = "DESC. SIN STOCK", C_BLUE_LIGHT
        else:
            status, bg = "Sin descuento", C_WHITE if ri % 2 == 0 else C_GRAY

        frequent = r.get("frequent", False)
        row = [ri - 1, "Sí" if frequent else "", r["title"], r["category"],
               "Sí" if in_stock else "No", status,
               r["original_price"], r["discounted_price"],
               r["savings_pct"], r["url"]]
        for ci, val in enumerate(row, start=1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", horizontal="center" if ci == 2 else "left")
            if ci in (7, 8) and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0'
            if ci == 9 and isinstance(val, (int, float)):
                cell.number_format = '0.0"%"'
            if ci == 6 and "DESCUENTO" in status:
                cell.font = Font(bold=True, color=C_ORANGE if in_stock else C_BLUE_DARK)
            if ci == 5 and not in_stock:
                cell.font = Font(bold=True, color=C_BLUE_DARK)

    # ── Sheet 4: Summary ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Resumen")
    summary_data = [
        ("Total productos analizados",      len(results)),
        ("Leídos correctamente",            len(results) - len(not_found)),
        ("Errores al leer",                 len(not_found)),
        ("Con descuento (con stock)",        len(disc_instock)),
        ("Con descuento (sin stock)",        len(disc_nostock)),
        ("Sin descuento",                   len(no_discount)),
        ("", ""),
        ("Fecha del reporte",               datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]

    _hcell(ws4, 1, 1, "Indicador", bg=C_GREEN_DARK)
    _hcell(ws4, 1, 2, "Valor",     bg=C_GREEN_DARK)
    ws4.column_dimensions["A"].width = 38
    ws4.column_dimensions["B"].width = 20

    for ri, (label, value) in enumerate(summary_data, start=2):
        ws4.cell(row=ri, column=1, value=label).font = Font(bold=bool(label))
        ws4.cell(row=ri, column=2, value=value)

    ws4.cell(row=4, column=1).font = Font(bold=True, color=C_ORANGE)
    ws4.cell(row=4, column=2).font = Font(bold=True, color=C_ORANGE)
    ws4.cell(row=5, column=1).font = Font(bold=True, color=C_BLUE_DARK)
    ws4.cell(row=5, column=2).font = Font(bold=True, color=C_BLUE_DARK)

    wb.save(filename)


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------

def send_email(filename: str, discounted: list[dict], total: int) -> None:
    if not ENABLE_EMAIL:
        print("\n--- Email desactivado (ENABLE_EMAIL = False) ---")
        return

    msg = MIMEMultipart()
    msg["From"]    = SENDER_EMAIL
    msg["To"]      = RECIPIENT_EMAIL
    msg["Subject"] = (
        f"Supermu Reporte {datetime.now().strftime('%Y-%m-%d')} "
        f"— {len(discounted)} descuento(s) encontrado(s)"
    )

    body_lines = [
        "<h2>Supermu — Reporte diario de descuentos</h2>",
        f"<p>Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>",
        f"<p>Productos analizados: <b>{total}</b> | Con descuento: <b>{len(discounted)}</b></p>",
    ]

    if discounted:
        top = sorted(discounted, key=lambda x: x["savings_pct"] or 0, reverse=True)
        body_lines.append("<h3>Productos con descuento (mayor a menor ahorro)</h3><ul>")
        for r in top:
            savings = f" — Ahorro: {r['savings_pct']}%" if r["savings_pct"] else ""
            link    = f" <a href='{r['url']}'>Ver</a>" if r["url"] else ""
            body_lines.append(
                f"<li><b>{r['title']}</b>: {fmt_cop(r['discounted_price'])} "
                f"(antes {fmt_cop(r['original_price'])}){savings}{link}</li>"
            )
        body_lines.append("</ul>")
    else:
        body_lines.append("<p>No se detectaron descuentos hoy.</p>")

    body_lines.append("<p><i>Reporte completo adjunto en Excel.</i></p>")
    msg.attach(MIMEText("\n".join(body_lines), "html"))

    with open(filename, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(filename)}"')
    msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.sendmail(SENDER_EMAIL, RECIPIENT_EMAIL, msg.as_string())
        print(f"  Email enviado a {RECIPIENT_EMAIL}")
    except Exception as e:
        print(f"  [ERROR] No se pudo enviar el email: {e}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("\nSupermu Discount Tracker")
    print("Fuente: /products.json (paginado)\n")

    all_results = scrape_all_products()

    total        = len(all_results)
    discounted   = [r for r in all_results if r["has_discount"]]
    disc_instock = [r for r in discounted if r.get("in_stock", True)]
    disc_nostock = [r for r in discounted if not r.get("in_stock", True)]
    not_found    = [r for r in all_results if not r["found"]]

    print(f"\n{'='*55}")
    print(f"  RESUMEN")
    print(f"{'='*55}")
    print(f"  Productos únicos:       {total}")
    print(f"  Con descuento + stock:  {len(disc_instock)}")
    print(f"  Con descuento sin stock:{len(disc_nostock)}")
    print(f"  Errores:                {len(not_found)}")

    if discounted:
        print(f"\n  --- Top descuentos ---")
        for r in sorted(discounted, key=lambda x: x["savings_pct"] or 0, reverse=True)[:10]:
            print(f"  {r['title'][:45]:<45} {r['discount_label']}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"supermu_descuentos_{timestamp}.xlsx"
    export_excel(all_results, filename)
    print(f"\n  Reporte guardado: {filename}")

    send_email(filename, discounted, total)


if __name__ == "__main__":
    main()
